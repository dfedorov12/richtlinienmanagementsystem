# -*- coding: utf-8 -*-
"""
Governance-Struktur aus der Zuständigkeiten-Mappe einlesen
==========================================================
Erzeugt den Datenteil von `js/govstruktur.js` aus der Excel-Mappe des
Corporate-Governance-Boards. Der Ansichtsteil der Datei (ab „Ansicht") bleibt
unangetastet – es wird nur der Kopf mit GOV_ARTEN / GOV_KATEGORIEN / GOV_STAND /
GOV_EINTRAEGE / GOV_WEITERE ersetzt.

Aufruf (im Projektverzeichnis, benötigt `python -m pip install openpyxl`):

    python scripts/govstruktur-import.py "C:/Pfad/CGB_Organisation_Zuständigkeiten_Nomenklatur.xlsx"

Ohne Pfad wird der Standardort auf dem Desktop versucht.

Aufbau der Mappe (Blatt 1, ab Zeile 6):
    A Grobfunktion · B Kategorie · C Zwischenüberschrift · D Titel
    E Verantwortung · F Dokumentname · G Version · H Datum · I Status
Zeilen ohne Verantwortung sind Zwischenüberschriften und werden übersprungen.
"""
import io
import json
import os
import re
import sys
import datetime

try:
    import openpyxl
except ImportError:
    sys.exit('Fehlt: openpyxl  →  python -m pip install openpyxl')

STANDARD = os.path.expanduser(
    r'~\OneDrive - dihag.com\Desktop\CGB_Organisation_Zuständigkeiten_Nomenklatur.xlsx')
ZIEL = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'js', 'govstruktur.js')

# Kategorien des Konzernregelwerk-Fundaments (Reihenfolge = Spalten der Übersicht)
KATEGORIEN = {
    '00_Allgemein': 'Allgemein',
    '01_Recht, Steuern, Datenschutz und Versicherungen': 'Recht / Steuern / Datenschutz / Versicherungen',
    '02_Compliance': 'Compliance',
    '03_Security/Cyber Security': 'Security / Cyber Security',
    '04_Finanzen/Rechnungswesen/Controlling/Einkauf': 'Finanzen / ReWe / Controlling / Einkauf',
    '05_Nachhaltigkeit, Arbeitssicherheit & Gesundheitsschutz': 'Nachhaltigkeit / Arbeitssicherheit & Gesundheitsschutz',
    '06_HR/Corporate Tranformation/Information Technology': 'HR / Corporate Transformation / IT',
}

ARTEN = [
    ('Handbuch', 'In sich abgeschlossenes Themengebiet (z. B. Code of Conduct)'),
    ('Policy', 'Strategischer Rahmen: Was ist das Ziel, der Grundsatz?'),
    ('Konzernrichtlinie', 'Operativer Rahmen: Wie handeln wir?'),
    ('Konzernfachregelung', 'Fachgerechte Ausführung'),
    ('Arbeits-/Prozessanweisung', 'Handlungsanleitung Schritt für Schritt'),
    ('Leitfaden', 'Handlungsempfehlungen'),
    ('Weitere', 'Vorlagen und Muster – außerhalb der Pyramide'),
]


def txt(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d.%m.%Y')
    return re.sub(r'\s+', ' ', str(v)).strip()


def _art(t):
    """Dokumentenart aus einem Text erkennen – '' wenn nichts eindeutig ist."""
    if 'muster' in t or 'template' in t: return 'Weitere'
    if 'verhaltenskodex' in t or 'code of conduct' in t or 'kodex' in t: return 'Handbuch'
    if 'handbuch' in t: return 'Handbuch'
    if 'leitfaden' in t: return 'Leitfaden'
    if 'anweisung' in t: return 'Arbeits-/Prozessanweisung'
    if 'fachregelung' in t: return 'Konzernfachregelung'
    if 'richtlinie' in t: return 'Konzernrichtlinie'
    if 'policy' in t or 'policies' in t or 'grundsatzerklärung' in t: return 'Policy'
    if 'anleitung' in t: return 'Leitfaden'
    if 'mindeststandard' in t or 'mindesstandard' in t: return 'Konzernfachregelung'
    return ''


def art_von(titel, ueberschrift):
    """Der Titel schlägt die Überschrift: Unter „…_Konzernrichtlinien und Policy"
       stehen beide Arten gemischt, die Überschrift führt dort in die Irre."""
    return _art(titel.lower()) or _art(ueberschrift.lower()) or 'Konzernrichtlinie'


def titel_von(t):
    """Führende Nummerierung und die schon in der Spalte stehende Art abschneiden
       (samt der Tippfehler, die in der Mappe vorkommen)."""
    t = re.sub(r'^\d+_\s*', '', t).strip()
    return re.sub(r'^(Konzern-?richtlinie|Konzerrichtlinie|Konzern-?fachregelung|Konzerfachregelung|Policy|Leitfaden)[\s_:]+',
                  '', t, flags=re.I).strip()


def status_von(roh):
    s = roh.lower()
    if not s: return 'offen'
    if 'final' in s or 'gültig' in s or 'veröffentlicht' in s: return 'gueltig'
    if 'offen' in s: return 'offen'
    return 'arbeit'      # in Prüfung, zu prüfen, in Erarbeitung, In Überarbeitung, Im Prozess


def einlesen(pfad):
    ws = openpyxl.load_workbook(pfad, data_only=True).worksheets[0]
    eintraege, weitere = [], []
    grob = kategorie = ueberschrift = ''
    for r in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        a, b, c, d, e, f, g, h, i = [txt(x) for x in r[:9]]
        if a:
            grob = a
        if grob == 'Konzernregelwerk':
            if b:
                kategorie, ueberschrift = b, ''
            if c:
                ueberschrift = c
            titel = d or c
            if not titel or titel.isdigit() or not e:
                continue                       # ohne Verantwortung: Zwischenüberschrift
            kat = KATEGORIEN.get(kategorie)
            if not kat:
                continue
            eintraege.append({
                'kategorie': kat, 'art': art_von(titel, ueberschrift), 'titel': titel_von(titel),
                'owner': e, 'status': status_von(i), 'statusRoh': i,
                'dokument': f, 'version': g, 'datum': h,
            })
        elif grob in ('Leitbild/Vision', 'Unternehmenspolitik', 'Kollektivrechtliche Regelungen'):
            titel = d or c or b or a
            if not titel or not (e or f):
                continue
            weitere.append({'bereich': grob, 'titel': re.sub(r'^\d+_\s*', '', titel),
                            'owner': e, 'status': status_von(i), 'statusRoh': i, 'dokument': f})
    return eintraege, weitere


def js(v):
    return json.dumps(v, ensure_ascii=False)


def kopf_bauen(eintraege, weitere, stand):
    kats = []
    for e in eintraege:
        if e['kategorie'] not in kats:
            kats.append(e['kategorie'])
    z = []
    z.append('/**')
    z.append(' * Reiter „Governance-Struktur"')
    z.append(' * ============================')
    z.append(' * Das Konzernregelwerk als Matrix: <b>Kategorie</b> (Spalten des')
    z.append(' * Konzernregelwerk-Fundaments) × <b>Dokumentenart</b> (Verbindlichkeitsebene der')
    z.append(' * Regelwerkspyramide), dazu Verantwortung und Stand je Regelung.')
    z.append(' *')
    z.append(' * Die Daten sind ein Auszug aus der Zuständigkeiten-Mappe des Corporate-Governance-')
    z.append(' * Boards (CGB_Organisation_Zuständigkeiten_Nomenklatur.xlsx, Stand ' + stand + ').')
    z.append(' * Sie stehen bewusst hier im Code und nicht in SharePoint: Es ist eine')
    z.append(' * Momentaufnahme der Planung, kein Live-Bestand. Ändert sich die Mappe, erzeugt')
    z.append(' * `python scripts/govstruktur-import.py <mappe.xlsx>` diesen Block neu.')
    z.append(' */')
    z.append('')
    z.append('/** Verbindlichkeitsebenen der Regelwerkspyramide – oben am verbindlichsten. */')
    z.append('const GOV_ARTEN = [')
    breite = max(len(k) for k, _ in ARTEN) + 3
    for key, erk in ARTEN:
        z.append('  { key: ' + (js(key) + ',').ljust(breite) + ' erklaerung: ' + js(erk) + ' },')
    z.append('];')
    z.append('')
    z.append('/** Kategorien des Konzernregelwerk-Fundaments (Reihenfolge wie in der Übersicht). */')
    z.append('const GOV_KATEGORIEN = [')
    for k in kats:
        z.append('  ' + js(k) + ',')
    z.append('];')
    z.append('')
    z.append('/** Stand der zugrunde liegenden Mappe. */')
    z.append('const GOV_STAND = ' + js(stand) + ';')
    z.append('')
    z.append('/** Ein Eintrag: { kategorie, art, titel, owner, status, … }.')
    z.append(" *  status: 'gueltig' (final abgelegt) | 'arbeit' (in Erarbeitung/Prüfung) | 'offen' (noch nicht begonnen).")
    z.append(' *  Titel und Verantwortung stehen unverändert so in der Mappe. */')
    z.append('const GOV_EINTRAEGE = [')
    for e in eintraege:
        felder = ['kategorie: ' + js(e['kategorie']), 'art: ' + js(e['art']), 'titel: ' + js(e['titel']),
                  'owner: ' + js(e['owner']), 'status: ' + js(e['status'])]
        felder += [f + ': ' + js(e[f]) for f in ('statusRoh', 'dokument', 'version', 'datum') if e.get(f)]
        z.append('  { ' + ', '.join(felder) + ' },')
    z.append('];')
    z.append('')
    z.append('/** Regelungen außerhalb des Konzernregelwerks – gleiche Mappe, eigene Grobfunktion. */')
    z.append('const GOV_WEITERE = [')
    for e in weitere:
        felder = ['bereich: ' + js(e['bereich']), 'titel: ' + js(e['titel']),
                  'owner: ' + js(e['owner']), 'status: ' + js(e['status'])]
        felder += [f + ': ' + js(e[f]) for f in ('statusRoh', 'dokument') if e.get(f)]
        z.append('  { ' + ', '.join(felder) + ' },')
    z.append('];')
    return '\n'.join(z) + '\n'


def main():
    pfad = sys.argv[1] if len(sys.argv) > 1 else STANDARD
    if not os.path.exists(pfad):
        sys.exit('Mappe nicht gefunden: ' + pfad)
    eintraege, weitere = einlesen(pfad)
    if not eintraege:
        sys.exit('Keine Einträge erkannt – Aufbau der Mappe geändert?')
    stand = datetime.datetime.fromtimestamp(os.path.getmtime(pfad)).strftime('%d.%m.%Y')

    alt = io.open(ZIEL, encoding='utf-8').read()
    marke = '\n/* ═══════════════════════════════════════════════════\n   Ansicht'
    if marke not in alt:
        sys.exit('Ansichtsteil in js/govstruktur.js nicht gefunden – Datei von Hand prüfen.')
    ansicht = alt[alt.index(marke):]
    io.open(ZIEL, 'w', encoding='utf-8', newline='\n').write(kopf_bauen(eintraege, weitere, stand) + ansicht)

    from collections import Counter
    print(f'{len(eintraege)} Regelungen, {len(weitere)} weitere · Stand {stand}')
    print('Je Art:      ', dict(Counter(e['art'] for e in eintraege)))
    print('Je Stand:    ', dict(Counter(e['status'] for e in eintraege)))
    print('Geschrieben: ', ZIEL)


if __name__ == '__main__':
    main()
